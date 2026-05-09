"""
=====================================================================
  LECTOR DE PRECIOS EN IMÁGENES — Visión Artificial con Claude
=====================================================================
  Lee los precios directamente de las fotos de posts de Instagram
  usando la API de Claude con visión artificial.

  Instalación:
    pip install anthropic requests pandas openpyxl pillow

  Uso:
    python leer_precios_imagenes.py --input ig_solo_ofertas.csv
    python leer_precios_imagenes.py --input ig_solo_ofertas.csv --limite 50

  Necesitás una API key de Anthropic:
    1. Ir a console.anthropic.com
    2. API Keys → Create Key
    3. Copiarla abajo en API_KEY
=====================================================================
"""

import argparse
import base64
import json
import logging
import os
import time
from pathlib import Path

import anthropic
import pandas as pd
import requests
from PIL import Image
from io import BytesIO

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("vision")

# ── CONFIGURACIÓN ─────────────────────────────────────────────────
# Pegá tu API key de console.anthropic.com
API_KEY = "TU_API_KEY_ACÁ"

# Pausa entre requests para no exceder límites
DELAY_ENTRE_IMAGENES = 1.5

# ── PROMPT PARA CLAUDE ────────────────────────────────────────────
PROMPT_VISION = """Sos un asistente de inteligencia comercial para un supermercado de Paraguay.

Analizá esta imagen de un post de Instagram de un supermercado competidor y extraé la siguiente información en formato JSON:

{
  "tiene_oferta": true/false,
  "productos": [
    {
      "nombre": "nombre del producto",
      "precio_gs": precio en guaraníes como número (sin símbolos),
      "precio_anterior_gs": precio anterior tachado si lo hay,
      "descuento_pct": porcentaje de descuento si está visible,
      "unidad": "kg/L/unidad/pack/etc",
      "vigencia": "fecha hasta cuando vale la oferta si se ve"
    }
  ],
  "tipo_oferta": "precio_especial/2x1/descuento_pct/precio_mayorista/otro",
  "supermercado": "nombre del supermercado si se ve en la imagen",
  "notas": "cualquier info relevante adicional"
}

Si la imagen no contiene precios o productos, devolvé:
{"tiene_oferta": false, "productos": [], "notas": "sin precios visibles"}

Respondé SOLO con el JSON, sin texto adicional ni backticks."""


def imagen_a_base64(url: str) -> tuple[str, str] | None:
    """Descarga imagen y la convierte a base64."""
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()

        # Verificar que es una imagen
        img = Image.open(BytesIO(resp.content))

        # Redimensionar si es muy grande (ahorra tokens)
        max_size = 1200
        if max(img.size) > max_size:
            img.thumbnail((max_size, max_size), Image.LANCZOS)

        # Convertir a JPEG base64
        buf = BytesIO()
        img.convert("RGB").save(buf, format="JPEG", quality=85)
        b64 = base64.standard_b64encode(buf.getvalue()).decode("utf-8")

        return b64, "image/jpeg"

    except Exception as e:
        log.warning(f"No se pudo procesar imagen {url[:60]}...: {e}")
        return None


def analizar_imagen(client: anthropic.Anthropic, imagen_b64: str, media_type: str) -> dict:
    """Envía imagen a Claude y obtiene datos de precios."""
    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1000,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": media_type,
                                "data": imagen_b64,
                            },
                        },
                        {
                            "type": "text",
                            "text": PROMPT_VISION
                        }
                    ],
                }
            ],
        )

        texto = response.content[0].text.strip()

        # Limpiar posibles backticks
        texto = texto.replace("```json", "").replace("```", "").strip()

        return json.loads(texto)

    except json.JSONDecodeError:
        log.warning("Respuesta no es JSON válido")
        return {"tiene_oferta": False, "productos": [], "notas": "error_parsing"}
    except Exception as e:
        log.error(f"Error en API Claude: {e}")
        return {"tiene_oferta": False, "productos": [], "notas": f"error_api: {str(e)}"}


def procesar_csv(input_csv: str, limite: int | None = None) -> pd.DataFrame:
    """Procesa el CSV de Instagram y agrega datos de precios de imágenes."""

    df = pd.read_csv(input_csv)
    log.info(f"Posts cargados: {len(df)}")

    # Filtrar solo cuentas reales
    cuentas_reales = [
        'supermercados_stock', 'supermas.py', 'comercialelcaciquepy',
        'boxmayoristapy', 'cadenareal', 'supermercados_superseis',
        'sancayetanocomercialpy', 'fortispy'
    ]
    nombres = {
        'supermercados_stock': 'Stock / S6',
        'supermas.py': 'SuperMás',
        'comercialelcaciquepy': 'El Cacique',
        'boxmayoristapy': 'Box Mayorista',
        'cadenareal': 'Cadena Real',
        'supermercados_superseis': 'Superseis',
        'sancayetanocomercialpy': 'San Cayetano',
        'fortispy': 'Fortis',
    }

    if 'cuenta' in df.columns:
        df = df[df['cuenta'].isin(cuentas_reales)].copy()
        df['competidor'] = df['cuenta'].map(nombres)

    if limite:
        df = df.head(limite)

    log.info(f"Posts a procesar: {len(df)}")

    # Verificar API key
    if API_KEY == "TU_API_KEY_ACÁ":
        log.error("❌ Configurá tu API key en el archivo antes de correr")
        log.error("   Ir a console.anthropic.com → API Keys → Create Key")
        return df

    client = anthropic.Anthropic(api_key=API_KEY)

    # Columnas nuevas
    resultados = []

    for i, row in enumerate(df.itertuples(), 1):
        url_imagen = getattr(row, 'url_imagen', '') or getattr(row, 'displayUrl', '')
        competidor = getattr(row, 'competidor', getattr(row, 'cuenta', ''))

        log.info(f"[{i}/{len(df)}] {competidor} — procesando imagen...")

        if not url_imagen or pd.isna(url_imagen):
            resultados.append({
                "vision_tiene_oferta": False,
                "vision_productos": "",
                "vision_precios": "",
                "vision_tipo_oferta": "",
                "vision_supermercado": "",
                "vision_vigencia": "",
                "vision_notas": "sin_url_imagen",
                "vision_procesado": False,
            })
            continue

        # Descargar y analizar imagen
        resultado_img = imagen_a_base64(str(url_imagen))

        if not resultado_img:
            resultados.append({
                "vision_tiene_oferta": False,
                "vision_productos": "",
                "vision_precios": "",
                "vision_tipo_oferta": "",
                "vision_supermercado": "",
                "vision_vigencia": "",
                "vision_notas": "error_descarga",
                "vision_procesado": False,
            })
            continue

        b64, media_type = resultado_img
        datos = analizar_imagen(client, b64, media_type)

        # Formatear productos y precios para Excel
        productos_str = ""
        precios_str = ""
        vigencia_str = ""

        if datos.get("productos"):
            prods = datos["productos"]
            productos_str = " | ".join(
                p.get("nombre", "") for p in prods if p.get("nombre")
            )
            precios_partes = []
            for p in prods:
                nombre = p.get("nombre", "")
                precio = p.get("precio_gs")
                anterior = p.get("precio_anterior_gs")
                desc = p.get("descuento_pct")
                if precio:
                    parte = f"{nombre}: ₲{precio:,.0f}".replace(",", ".")
                    if anterior:
                        parte += f" (antes ₲{anterior:,.0f})".replace(",",".")
                    if desc:
                        parte += f" -{desc}%"
                    precios_partes.append(parte)
                vigencia_str = p.get("vigencia", "") or vigencia_str

            precios_str = " | ".join(precios_partes)

        resultados.append({
            "vision_tiene_oferta": datos.get("tiene_oferta", False),
            "vision_productos": productos_str,
            "vision_precios": precios_str,
            "vision_tipo_oferta": datos.get("tipo_oferta", ""),
            "vision_supermercado": datos.get("supermercado", ""),
            "vision_vigencia": vigencia_str,
            "vision_notas": datos.get("notas", ""),
            "vision_procesado": True,
        })

        log.info(f"  ✓ {'Oferta detectada: ' + precios_str[:60] if precios_str else 'Sin precios visibles'}")
        time.sleep(DELAY_ENTRE_IMAGENES)

    # Agregar columnas al DataFrame
    df_resultados = pd.DataFrame(resultados)
    df_final = pd.concat([df.reset_index(drop=True), df_resultados], axis=1)

    return df_final


def guardar_excel(df: pd.DataFrame, output: str):
    """Guarda resultados en Excel formateado."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def fill(hex): return PatternFill("solid", start_color=hex, fgColor=hex)
    def fnt(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size, name="Arial")
    def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
    def brd():
        s = Side(style="thin", color="DDDDDD")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Precios Detectados"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Título
    ws.merge_cells("A1:L1")
    ws["A1"] = "PRECIOS DETECTADOS EN IMÁGENES — Visión Artificial"
    ws["A1"].font = fnt(True, "FFFFFF", 13)
    ws["A1"].fill = fill("1A237E")
    ws["A1"].alignment = ctr()
    ws.row_dimensions[1].height = 34

    # Columnas a mostrar
    cols_mostrar = {
        'competidor': ('Competidor', 15),
        'fecha': ('Fecha', 12),
        'descripcion': ('Descripción', 40),
        'likes': ('Likes', 8),
        'vision_tiene_oferta': ('Tiene Oferta', 12),
        'vision_productos': ('Productos (IA)', 35),
        'vision_precios': ('Precios (IA)', 40),
        'vision_tipo_oferta': ('Tipo', 14),
        'vision_vigencia': ('Vigencia', 14),
        'vision_supermercado': ('Supermercado', 15),
        'url_post': ('URL Post', 40),
        'url_imagen': ('URL Imagen', 40),
    }

    cols_validas = {k: v for k, v in cols_mostrar.items() if k in df.columns}

    for i, (col, (header, width)) in enumerate(cols_validas.items(), 1):
        c = ws.cell(row=2, column=i, value=header)
        c.font = fnt(True, "FFFFFF", 9)
        c.fill = fill("1565C0")
        c.alignment = ctr()
        c.border = brd()
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[2].height = 28

    colores_cuenta = {
        'Stock / S6': "E3F2FD", 'SuperMás': "E8F5E9",
        'El Cacique': "FFFDE7", 'Box Mayorista': "FFF3E0",
        'Cadena Real': "FCE4EC", 'Superseis': "F3E5F5",
        'San Cayetano': "E0F7FA", 'Fortis': "F5F5F5",
    }

    for i, row in enumerate(df.itertuples(), 3):
        comp = getattr(row, 'competidor', '')
        bg = colores_cuenta.get(comp, "FFFFFF")
        if i % 2 == 0:
            bg = bg
        else:
            bg = "FFFFFF"

        for j, col in enumerate(cols_validas.keys(), 1):
            val = getattr(row, col, '')
            if isinstance(val, float) and pd.isna(val):
                val = ''
            c = ws.cell(row=i, column=j, value=str(val)[:200] if val else '')
            c.font = fnt(size=9, bold=(j == 1))
            c.border = brd()
            c.alignment = lft() if j in (3, 6, 7) else ctr()
            c.fill = fill(bg)

            # Highlight ofertas detectadas
            if col == 'vision_tiene_oferta':
                if val is True or val == 'True':
                    c.font = fnt(True, "2E7D32", 9)
                    c.fill = fill("E8F5E9")
            if col == 'vision_precios' and val:
                c.font = fnt(True, "1565C0", 9)

    wb.save(output)
    log.info(f"✅ Excel guardado: {output}")

    # Estadísticas
    if 'vision_procesado' in df.columns:
        procesados = df['vision_procesado'].sum()
        con_oferta = df['vision_tiene_oferta'].sum() if 'vision_tiene_oferta' in df.columns else 0
        log.info(f"\n📊 Resumen:")
        log.info(f"   Imágenes procesadas: {procesados}")
        log.info(f"   Con ofertas detectadas: {con_oferta}")
        log.info(f"   Sin precio visible: {procesados - con_oferta}")


def main():
    parser = argparse.ArgumentParser(description="Lector de precios en imágenes de Instagram")
    parser.add_argument("--input", default="ig_solo_ofertas.csv", help="CSV de posts de Instagram")
    parser.add_argument("--output", default="precios_detectados.xlsx", help="Excel de salida")
    parser.add_argument("--limite", type=int, default=None, help="Limitar a N posts (para prueba)")
    args = parser.parse_args()

    if not Path(args.input).exists():
        log.error(f"Archivo no encontrado: {args.input}")
        log.info("Asegurate de tener el archivo ig_solo_ofertas.csv en la misma carpeta")
        return

    log.info("=" * 60)
    log.info("LECTOR DE PRECIOS EN IMÁGENES — Visión Artificial")
    log.info("=" * 60)

    df = procesar_csv(args.input, args.limite)
    guardar_excel(df, args.output)

    log.info(f"\n✅ COMPLETADO — Resultados en: {args.output}")


if __name__ == "__main__":
    main()
